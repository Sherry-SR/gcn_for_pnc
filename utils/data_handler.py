import importlib

import os
import numpy as np
import pandas as pd
from scipy.sparse import coo_matrix

import itertools
from operator import itemgetter
from openpyxl import load_workbook
import pickle
from shutil import copyfile

import torch
from torch_geometric.utils import dense_to_sparse
from torch_geometric.data import Data, InMemoryDataset, DataLoader

def train_val_test_split(path, output, train_ratio, val_ratio, test_ratio):
    if os.path.splitext(path)[1] == '.txt':
        filelist = np.loadtxt(path, dtype=str)
    else:
        filelist = os.listdir(path)
    np.random.shuffle(filelist)

    sum_ratio = train_ratio + val_ratio + test_ratio
    train_ratio = train_ratio / sum_ratio
    val_ratio = val_ratio / sum_ratio
    test_ratio = test_ratio / sum_ratio

    train_list = filelist[0:int(len(filelist) * train_ratio)]
    val_list = filelist[int(len(filelist) * train_ratio):-int(len(filelist) * test_ratio)]
    test_list = filelist[-int(len(filelist) * test_ratio):]

    print('train:', len(train_list), 'val:', len(val_list), 'test:', len(test_list))
    np.savetxt(os.path.join(output,'train_list.txt'), train_list, fmt='%s', delimiter='\n')
    np.savetxt(os.path.join(output,'val_list.txt'), val_list, fmt='%s', delimiter='\n')
    np.savetxt(os.path.join(output,'test_list.txt'), test_list, fmt='%s', delimiter='\n')
    print('filelist saved to:', output)

#train_val_test_split('/home/sherry/Dropbox/PhD/Data/ABIDE/abide_qc.txt', '/home/sherry/Dropbox/PhD/Data/ABIDE/abide_exp01', 0.7, 0.2, 0.1)

def cross_validation_split(path, output, num_fold, sel_fold = 0, shuffle = True):
    if os.path.splitext(path)[1] == '.txt':
        filelist = np.loadtxt(path, dtype=str)
    else:
        filelist = os.listdir(path)
    if shuffle:
        np.random.shuffle(filelist)
        np.savetxt(os.path.join(output,'all_list_shuffled.txt'), filelist, fmt='%s', delimiter='\n')

    ind0 = int(len(filelist) / num_fold) * sel_fold
    ind1 = int(len(filelist) / num_fold) * (sel_fold + 1)

    train_list = np.concatenate((filelist[ind1:], filelist[0:ind0]))
    val_list = filelist[ind0:ind1]
    test_list = val_list

    print('train:', len(train_list), 'val:', len(val_list), 'test:', len(test_list))
    np.savetxt(os.path.join(output,'train_list.txt'), train_list, fmt='%s', delimiter='\n')
    np.savetxt(os.path.join(output,'val_list.txt'), val_list, fmt='%s', delimiter='\n')
    np.savetxt(os.path.join(output,'test_list.txt'), test_list, fmt='%s', delimiter='\n')
    print('filelist saved to:', output)

#cross_validation_split('/home/sherry/Dropbox/PhD/Data/ABIDE/abide_exp01/all_list_shuffled.txt', '/home/sherry/Dropbox/PhD/Data/ABIDE/abide_exp01', 10, 9, False)

def arrange_data(path, output):
    subpaths = [f.path for f in os.scandir(path) if f.is_dir()]
    for subpath in subpaths:
        filelist = os.listdir(subpath)
        for filename in filelist:
            subj, ext = os.path.splitext(filename)
            out_subpath = os.path.join(output, subj)
            if not os.path.exists(out_subpath):
                os.mkdir(out_subpath)
            out_filename = os.path.basename(path)+'_'+os.path.basename(subpath)+'_matrix'+ext
            copyfile(os.path.join(subpath, filename), os.path.join(out_subpath, out_filename))

#arrange_data('/home/sherry/Dropbox/PhD/Data/ABIDE/raw_data/dos160', '/home/sherry/Dropbox/PhD/Data/ABIDE/ABIDE_Connectomes')

def read_xlsx(path):
    workbook = load_workbook(path)
    sheet = workbook[workbook.sheetnames[0]]
    data = sheet.values
    cols = next(data)[0:]
    data = list(data)
    data = (itertools.islice(r, 0, None) for r in data)
    df = pd.DataFrame(data, columns = cols)
    return df

class ABIDESet(InMemoryDataset):
    def __init__(self, sub_list, output, root, path_data, path_label, target_name = None, feature_mask = None, **kwargs):
        self.path = [path_data, path_label]
        self.output = output
        if sub_list is None:
            sub_list = os.listdir(path_data)
        self.sub_list = sub_list
        self.target_name = target_name
        self.feature_mask = feature_mask
        super(ABIDESet, self).__init__(root)
        self.data, self.slices = torch.load(self.processed_paths[0])

    @property
    def raw_file_names(self):
        return ['abide_raw.pkl', 'abide_raw_info.txt']

    @property
    def processed_file_names(self):
        return [self.output]

    def download(self):
        assert len(self.path) == 2
        path_data = self.path[0]
        path_label = self.path[1]
        
        labels = read_xlsx(path_label)
        labels = labels.astype({'subject': 'str'})
        labels['SITE_ID'], uniques = pd.factorize(labels['SITE_ID'])
        labels['DX_GROUP'] = 2 - labels['DX_GROUP']
        labels['SEX'] = labels['SEX'] - 1

        labels = itemgetter('SITE_ID','DX_GROUP','DSM_IV_TR','AGE_AT_SCAN','SEX')(labels.set_index('subject').to_dict())
        
        subjlist = os.listdir(path_data)
        filelist = [fname for fname in os.listdir(os.path.join(path_data, subjlist[0])) if 'matrix' in fname]

        with open(os.path.join(self.raw_dir, 'abide_raw_info.txt'), 'w') as f:
            print('Label info:', file = f)
            print('All labels:', 'SITE_ID','DX_GROUP','DSM_IV_TR','AGE_AT_SCAN','SEX')
            print('Site labels (0-n):', uniques.values, file = f)
            print('DX_GROUP (0/1):', 'control, autism', file = f)
            print('DSM_IV_TR (0-n):', 'control, autism, aspergers, PDD-NOS, aspergers or PDD-NOS', file = f)
            print('SEX (0/1):', 'M, F', file = f)
            print('\n', file = f)
            print('Features:', file = f)
            print(filelist, sep='\n', file = f)
            print('\n', file = f)
            print('Saved subjects:', file = f)
            print(subjlist, sep='\n', file = f)
            print('\n', file = f)

        dataset = {}
        for subj in subjlist:
            print('downloading', subj, '...')
            features = []
            for file in filelist:
                filepath = os.path.join(path_data, subj, file)
                # origianl value (-1 ~ 1), adjust value of the matrix to 0 ~ 2
                matrix = torch.tensor(np.loadtxt(filepath), dtype=torch.float) + 1
                features.append(matrix)
            
            y = {'SITE_ID': labels[0][subj], 'DX_GROUP': labels[1][subj], 'DSM_IV_TR': labels[2][subj],
                    'AGE_AT_SCAN': labels[3][subj], 'SEX': labels[4][subj]}
            x = torch.ones([matrix.shape[0], 1], dtype=torch.float)
            data = Data(x = x, y = y)
            data.features = features           
            dataset[subj] = data
    
        with open(os.path.join(self.raw_dir, 'abide_raw.pkl'), 'wb') as f:
            pickle.dump(dataset, f)
            print('ABIDE dataset saved to path:', self.raw_dir)
        
    def process(self):
        with open(os.path.join(self.raw_dir, 'abide_raw.pkl'), 'rb') as f:
            dataset = pickle.load(f)

        dataset_list = []
        
        sub_list = np.loadtxt(self.sub_list, dtype = str, delimiter = '\n')

        for subj in sub_list:
            data = dataset[subj]
            if self.target_name is not None:
                data.y = data.y[self.target_name]
            if self.feature_mask is not None:
                data.features = [data.features[i] for i in self.feature_mask]

            edge_index, _ = dense_to_sparse(torch.ones(data.features[0].shape, dtype=torch.float))
            edge_attr = []
            for feature in data.features:
                edge_attr.append(feature[edge_index[0], edge_index[1]])
            data.edge_index = edge_index
            data.edge_attr = torch.stack(edge_attr, dim = -1)
            data.features = torch.stack(data.features, dim = -1)
            data.features = torch.unsqueeze(data.features, dim = 0)
            dataset_list.append(data)
            
        self.data, self.slices = self.collate(dataset_list)
        torch.save((self.data, self.slices), self.processed_paths[0])
        print('Processed dataset saved as', self.processed_paths[0])

    def __repr__(self):
        return '{}()'.format(self.__class__.__name__)


class PNCEnrichedSet(InMemoryDataset):
    def __init__(self, sub_list, output, root, path_data, path_label, target_name = None, feature_mask = None, **kwargs):
        self.path = [path_data, path_label]
        self.output = output
        if sub_list is None:
            sub_list = os.listdir(path_data)
        self.sub_list = sub_list
        self.target_name = target_name
        self.feature_mask = feature_mask
        super(PNCEnrichedSet, self).__init__(root)
        self.data, self.slices = torch.load(self.processed_paths[0])

    @property
    def raw_file_names(self):
        return ['pnc_features_raw.pkl']

    @property
    def processed_file_names(self):
        return [self.output]

    def download(self):
        assert len(self.path) == 2
        path_data = self.path[0]
        path_label = self.path[1]
        
        labels = read_xlsx(path_label)
        labels['Sex'], uniques = pd.factorize(labels['Sex'])
        labels = itemgetter('ScanAgeYears','Sex')(labels.set_index('Subject').to_dict())
        
        subjlist = os.listdir(path_data)
        #filelist = [fname for fname in os.listdir(os.path.join(path_data, subjlist[0])) if 'matrix' in fname]
        filelist = ['fdt_network_matrix', 'fdt_network_matrix_lengths',
                    'Enriched_mean_matrix_0', 'Enriched_variance_matrix_0',
                    'Enriched_mean_matrix_1', 'Enriched_variance_matrix_1',
                    'Enriched_mean_matrix_2', 'Enriched_variance_matrix_2',
                    'Enriched_mean_matrix_3', 'Enriched_variance_matrix_3']

        with open(os.path.join(self.raw_dir, 'pnc_enriched_raw_info.txt'), 'w') as f:
            print('Label info:', file = f)
            print('All labels:', 'ScanAgeYears','Sex')
            print('Sex labels (0/1):', uniques.values, file = f)
            print('\n', file = f)
            print('Enriched features:', file = f)
            print(filelist, sep='\n', file = f)
            print('\n', file = f)
            print('Saved subjects:', file = f)
            print(subjlist, sep='\n', file = f)
            print('\n', file = f)

        dataset = {}
        for subj in subjlist:
            print('downloading', subj, '...')
            features = []
            for file in filelist:
                filepath = os.path.join(path_data, subj, file)
                matrix = torch.tensor(np.loadtxt(filepath), dtype=torch.float)
                features.append(matrix)
            y = {'ScanAgeYears': labels[0][subj], 'Sex': labels[1][subj]}
            x = torch.ones([matrix.shape[0], 1], dtype=torch.float) 
            data = Data(x = x, y = y)
            data.features = features           
            dataset[subj] = data
    
        with open(os.path.join(self.raw_dir, 'pnc_features_raw.pkl'), 'wb') as f:
            pickle.dump(dataset, f)
            print('PNC dataset saved to path:', self.raw_dir)
        
    def process(self):
        with open(os.path.join(self.raw_dir, 'pnc_features_raw.pkl'), 'rb') as f:
            dataset = pickle.load(f)

        dataset_list = []
        
        sub_list = np.loadtxt(self.sub_list, dtype = str, delimiter = '\n')

        for subj in sub_list:
            data = dataset[subj]
            if self.target_name is not None:
                data.y = data.y[self.target_name]
            if self.feature_mask is not None:
                data.features = [data.features[i] for i in self.feature_mask]
                
            edge_index, _ = dense_to_sparse(torch.ones(data.features[0].shape, dtype=torch.float))
            edge_attr = []
            for feature in data.features:
                edge_attr.append(feature[edge_index[0], edge_index[1]])
            data.edge_index = edge_index
            data.edge_attr = torch.stack(edge_attr, dim = -1)
            data.features = torch.stack(data.features, dim = -1)
            data.features = torch.unsqueeze(data.features, dim = 0)
            dataset_list.append(data)
            
        self.data, self.slices = self.collate(dataset_list)
        torch.save((self.data, self.slices), self.processed_paths[0])
        print('Processed dataset saved as', self.processed_paths[0])

    def __repr__(self):
        return '{}()'.format(self.__class__.__name__)


def get_data_loaders(config):
    assert 'loaders' in config, 'Could not find loaders configuration'
    loaders_config = config['loaders']
    class_name = loaders_config.pop('name')
    train_list = loaders_config.pop('train_list')
    val_list = loaders_config.pop('val_list')
    test_list = loaders_config.pop('test_list')
    output_train = loaders_config.pop('output_train')
    output_val = loaders_config.pop('output_val')
    output_test = loaders_config.pop('output_test')
    batch_size = loaders_config.pop('batch_size')

    m = importlib.import_module('utils.data_handler')
    clazz = getattr(m, class_name)

    return {
        'train': DataLoader(clazz(train_list, output_train, **loaders_config), batch_size=batch_size, shuffle=True),
        'val': DataLoader(clazz(val_list, output_val, **loaders_config), batch_size=batch_size, shuffle=True),
        'test': DataLoader(clazz(test_list, output_test, **loaders_config), batch_size=batch_size, shuffle=True)
        }